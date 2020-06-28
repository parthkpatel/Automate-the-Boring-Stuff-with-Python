#! python3
# send_dues_reminders.py - Sends emails based on payment status in spreadsheet.

import openpyxl, smtplib, pyinputplus as pyip

# Open the spreadsheet and get the latest dues status.

wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb['Sheet1']
last_col = sheet.max_column
latest_month = sheet.cell(row=1, column=last_col).value

# Check each member's payment status.
unpaid_members = {}
for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=last_col).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaid_members[name] = email

# Log in to email account.
email = pyip.inputEmail(prompt='Please enter your gmail email address:\n')
password = pyip.inputStr(prompt='Please enter your email password:\n')
smtp_obj = smtplib.SMTP('smtp.gmail.com', 587)
smtp_obj.ehlo()
smtp_obj.starttls()
smtp_obj.login(email, password)

# Send out reminder emails.
for name, email in unpaid_members.items():
    body = f"Subject: {latest_month} dues unpaid.\nDear {name},\nRecords show that you have not " \
           f"paid dues for {latest_month}. Please make this payment as soon as possible. Thank you!"
    print('Sending email to %s...' % email)
    send_mail_status = smtp_obj.sendmail('my_email_address@example.com', email, body)

    if send_mail_status != {}:
        print('There was a problem sending email to %s: %s' % (email, send_mail_status))
smtp_obj.quit()
