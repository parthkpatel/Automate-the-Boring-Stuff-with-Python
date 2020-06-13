# A program that creates an nxn multiplication table and saves it in an Excel spreadsheet

import pyinputplus as pyip, openpyxl
from openpyxl.styles import Font


def create_multiplication_table(n):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = f'{n}x{n} Multiplication Table'
    bold_font_obj = Font(bold=True)

    for i in range(1, n+1):
        # Row headers and font
        sheet.cell(row=i+1, column=1).value = i
        sheet.cell(row=i + 1, column=1).font = bold_font_obj

        # Column headers and font
        sheet.cell(row=1, column=i+1).value = i
        sheet.cell(row=1, column=i+1).font = bold_font_obj

    # Multiplication table creation
    for rowNum in range(1, n+1):
        for colNum in range(1, n+1):
            sheet.cell(row=rowNum+1, column=colNum+1).value = rowNum*colNum

    wb.save('multiplication_table.xlsx')
    wb.close()


if __name__ == '__main__':
    number = pyip.inputInt(prompt='Input the integer you want a multiplication table for:\n', min=2)
    create_multiplication_table(number)
